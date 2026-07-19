import json, re, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Parse raw snapshot file ───────────────────────────────────────────────────
with open(r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_raw2.txt', encoding='utf-8') as f:
    content = f.read()

m = re.search(r'generic \[ref=\w+\]: (\"\{.+)', content)
if not m:
    raise RuntimeError("JSON snapshot not found")
quoted = m.group(1).rstrip()
if not quoted.endswith('"'):
    quoted += '"'
inner = json.loads(quoted)
data  = json.loads(inner)
print(f"Total cases: {data['size']}")

# Save raw JSON
with open(r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases2.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2)

# ── Helpers ───────────────────────────────────────────────────────────────────
def clean(html):
    if not html:
        return ''
    text = re.sub(r'<br\s*/?>', '\n', html)
    text = re.sub(r'<li>',       '\n- ', text)
    text = re.sub(r'<[^>]+>',   '', text)
    return re.sub(r'\n{3,}', '\n\n', text).strip()

# ── Build rows (one per step) ─────────────────────────────────────────────────
rows = []
for c in data['cases']:
    steps     = c.get('custom_steps_separated') or []
    preconds  = clean(c.get('custom_preconds') or '')
    case_id   = 'C' + str(c['id'])
    title     = c['title']
    if not steps:
        rows.append({
            'Case ID':          case_id,
            'Title':            title,
            'Pre-conditions':   preconds,
            'Step #':           '',
            'Step Description': '',
            'Expected Result':  '',
        })
    for i, s in enumerate(steps, 1):
        rows.append({
            'Case ID':          case_id,
            'Title':            title,
            'Pre-conditions':   preconds if i == 1 else '',
            'Step #':           i,
            'Step Description': clean(s.get('content', '')),
            'Expected Result':  clean(s.get('expected', '') or ''),
        })

# ── CSV ───────────────────────────────────────────────────────────────────────
csv_path = r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases2.csv'
with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
print(f"CSV  → {csv_path}  ({len(rows)} rows)")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Test Cases'

headers = list(rows[0].keys())
thin_side  = Side(style='thin')
border     = Border(thin_side, thin_side, thin_side, thin_side)
wrap_align = Alignment(wrap_text=True, vertical='top')

# Header row
for col, hdr in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font      = Font(bold=True, color='FFFFFF')
    cell.fill      = PatternFill('solid', fgColor='1F4E79')
    cell.border    = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data rows with alternating colour per case
fill_even = PatternFill('solid', fgColor='DCE6F1')
fill_odd  = PatternFill('solid', fgColor='FFFFFF')
prev_case    = None
colour_toggle = False

for row_idx, row in enumerate(rows, 2):
    if row['Case ID'] != prev_case:
        colour_toggle = not colour_toggle
        prev_case = row['Case ID']
    row_fill = fill_even if colour_toggle else fill_odd
    for col, key in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=row[key])
        cell.fill      = row_fill
        cell.border    = border
        cell.alignment = wrap_align

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 65
ws.column_dimensions['F'].width = 65

ws.freeze_panes  = 'A2'
ws.auto_filter.ref = ws.dimensions

xlsx_path = r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases2.xlsx'
wb.save(xlsx_path)
print(f"Excel → {xlsx_path}  ({len(rows)} rows)")

# ── Plain text ────────────────────────────────────────────────────────────────
lines = ["DUNE TEST CATALOGUE — Section 8505887", "=" * 70]
for idx, c in enumerate(data['cases'], 1):
    steps    = c.get('custom_steps_separated') or []
    preconds = clean(c.get('custom_preconds') or '')
    lines.append(f"\nTest Case {idx} of {data['size']}")
    lines.append(f"ID      : C{c['id']}")
    lines.append(f"Title   : {c['title']}")
    lines.append("-" * 70)
    if preconds:
        lines.append("Pre-conditions:")
        for ln in preconds.splitlines():
            if ln.strip():
                lines.append(f"  {ln.strip()}")
        lines.append("")
    if steps:
        lines.append("Steps:")
        for i, s in enumerate(steps, 1):
            content  = clean(s.get('content', ''))
            expected = clean(s.get('expected', '') or '')
            lines.append(f"\n  Step {i}:")
            for ln in content.splitlines():
                if ln.strip():
                    lines.append(f"    {ln.strip()}")
            if expected:
                lines.append("    Expected Result:")
                for ln in expected.splitlines():
                    if ln.strip():
                        lines.append(f"      {ln.strip()}")
    lines.append("\n" + "=" * 70)

txt_path = r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases2.txt'
with open(txt_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print(f"Text → {txt_path}")
