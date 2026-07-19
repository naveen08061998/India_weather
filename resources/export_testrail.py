import json, re, csv

def clean(html):
    if not html:
        return ''
    text = re.sub(r'<br\s*/?>', '\n', html)
    text = re.sub(r'<li>', '\n- ', text)
    text = re.sub(r'<[^>]+>', '', text)
    return re.sub(r'\n{3,}', '\n\n', text).strip()

with open(r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases.json', encoding='utf-8') as f:
    data = json.load(f)

rows = []
for c in data['cases']:
    steps = c.get('custom_steps_separated') or []
    preconds = clean(c.get('custom_preconds') or '')
    case_id = 'C' + str(c['id'])
    title   = c['title']
    for i, s in enumerate(steps, 1):
        rows.append({
            'Case ID':          case_id,
            'Title':            title,
            'Pre-conditions':   preconds if i == 1 else '',
            'Step #':           i,
            'Step Description': clean(s.get('content', '')),
            'Expected Result':  clean(s.get('expected', '') or ''),
        })

# ── CSV ────────────────────────────────────────────────────────────────────────
csv_path = r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases.csv'
with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
print(f'CSV  → {csv_path}  ({len(rows)} rows)')

# ── Excel ──────────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Test Cases'

headers = list(rows[0].keys())

# Header row styling
header_font  = Font(bold=True, color='FFFFFF')
header_fill  = PatternFill('solid', fgColor='1F4E79')
thin_side    = Side(style='thin')
border       = Border(thin_side, thin_side, thin_side, thin_side)
wrap_align   = Alignment(wrap_text=True, vertical='top')

for col, hdr in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font  = header_font
    cell.fill  = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Alternating row colours
fill_even = PatternFill('solid', fgColor='DCE6F1')
fill_odd  = PatternFill('solid', fgColor='FFFFFF')

prev_case = None
colour_toggle = False
for row_idx, row in enumerate(rows, 2):
    if row['Case ID'] != prev_case:
        colour_toggle = not colour_toggle
        prev_case = row['Case ID']
    row_fill = fill_even if colour_toggle else fill_odd

    for col, key in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=row[key])
        cell.fill      = row_fill
        cell.border    = border
        cell.alignment = wrap_align

# Column widths
ws.column_dimensions['A'].width = 14   # Case ID
ws.column_dimensions['B'].width = 45   # Title
ws.column_dimensions['C'].width = 35   # Pre-conditions
ws.column_dimensions['D'].width = 8    # Step #
ws.column_dimensions['E'].width = 60   # Step Description
ws.column_dimensions['F'].width = 60   # Expected Result

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

xlsx_path = r'C:\Users\ReddyA41\Desktop\AgenticAgent\resources\testrail_cases.xlsx'
wb.save(xlsx_path)
print(f'Excel → {xlsx_path}  ({len(rows)} rows)')
